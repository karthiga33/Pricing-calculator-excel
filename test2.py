import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
import requests
import json
from datetime import datetime
import os
import logging
from typing import Dict, List
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"
GROQ_MODEL = "llama-3.1-8b-instant"

# Get API key from Streamlit secrets, .env file, config.py, or environment variable
GROQ_API_KEY = ""
try:
    import streamlit as st
    if hasattr(st, 'secrets') and "GROQ_API_KEY" in st.secrets:
        GROQ_API_KEY = st.secrets["GROQ_API_KEY"]
        logger.info("Loaded API key from Streamlit secrets")
except Exception as e:
    logger.warning(f"Could not load from Streamlit secrets: {e}")

if not GROQ_API_KEY:
    try:
        from config import GROQ_API_KEY
        logger.info("Loaded API key from config.py")
    except ImportError:
        GROQ_API_KEY = os.getenv("GROQ_API_KEY", "")
        if GROQ_API_KEY:
            logger.info("Loaded API key from environment variable")
        else:
            logger.error("No API key found in any source!")

def call_groq(prompt: str, max_tokens: int = 500) -> str:
    """Call Groq API (Free)"""
    if not GROQ_API_KEY:
        logger.error("GROQ_API_KEY not set")
        return ""
    
    try:
        logger.info(f"Calling Groq API with max_tokens={max_tokens}")
        response = requests.post(GROQ_API_URL, 
            headers={
                "Authorization": f"Bearer {GROQ_API_KEY}",
                "Content-Type": "application/json"
            },
            json={
                "model": GROQ_MODEL,
                "messages": [{"role": "user", "content": prompt}],
                "max_tokens": max_tokens,
                "temperature": 0.5
            },
            timeout=30
        )
        logger.info(f"Groq API response status: {response.status_code}")
        if response.status_code == 200:
            result = response.json()['choices'][0]['message']['content']
            logger.info(f"Groq API success, response length: {len(result)}")
            return result
        else:
            logger.error(f"Groq API error: {response.status_code} - {response.text}")
            return ""
    except Exception as e:
        logger.error(f"Groq API connection error: {e}")
        return ""

class CostReportAgent:
    def __init__(self, default_usd_to_inr: float, default_region: str = "US East (N. Virginia)"):
        self.usd_to_inr = default_usd_to_inr
        self.default_region = default_region
        logger.info("CostReportAgent initialized with USD to INR rate: %.2f", self.usd_to_inr)

    def extract_ec2_specs(self, instance_types: List[str]) -> Dict:
        if not instance_types:
            return {}

        prompt = f"""You are an AWS EC2 specifications expert. Provide EXACT official AWS specifications.

Instance types to lookup: {', '.join(instance_types)}

CRITICAL RULES - Follow AWS official patterns:

1. SIZE PATTERNS (EXACT):
   - nano: 2 vCPUs, 0.5 GiB
   - micro: 1-2 vCPUs, 1 GiB
   - small: 1-2 vCPUs, 2 GiB
   - medium: 2 vCPUs, 4 GiB
   - large: 2 vCPUs, 8 GiB
   - xlarge: 4 vCPUs, 16 GiB
   - 2xlarge: 8 vCPUs, 32 GiB
   - 4xlarge: 16 vCPUs, 64 GiB

2. FAMILY PATTERNS:
   T-series (Burstable): t2, t3, t3a, t4g
   - t3.medium = 2 vCPUs, 4 GiB
   - t3a.medium = 2 vCPUs, 4 GiB (AMD variant)
   - t3.large = 2 vCPUs, 8 GiB
   
   M-series (General Purpose): m5, m6a, m6i, m6g
   - m6a.large = 2 vCPUs, 8 GiB
   - m6a.xlarge = 4 vCPUs, 16 GiB
   - m5.2xlarge = 8 vCPUs, 32 GiB
   
   C-series (Compute): c5, c6a, c6i
   - c5.large = 2 vCPUs, 4 GiB (HALF memory of M-series)
   - c5.xlarge = 4 vCPUs, 8 GiB
   
   R-series (Memory): r5, r6a, r6i
   - r5.large = 2 vCPUs, 16 GiB (DOUBLE memory of M-series)
   - r5.xlarge = 4 vCPUs, 32 GiB

3. MEMORY FORMULA:
   - T/M-series: large=8GB, xlarge=16GB, 2xlarge=32GB
   - C-series: large=4GB, xlarge=8GB, 2xlarge=16GB (half of M)
   - R-series: large=16GB, xlarge=32GB, 2xlarge=64GB (double of M)

Return ONLY valid JSON (no markdown, no text):
{{
  "t3a.medium": {{"vCPUs": 2, "MemoryGiB": 4}},
  "m6a.large": {{"vCPUs": 2, "MemoryGiB": 8}},
  "c5.xlarge": {{"vCPUs": 4, "MemoryGiB": 8}}
}}

JSON:"""

        try:
            response = call_groq(prompt, max_tokens=600)
            logger.info(f"EC2 specs raw response: {response[:200]}")
            
            # Try to extract JSON from response
            json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response, re.DOTALL)
            if json_match:
                json_str = json_match.group()
                ec2_specs = json.loads(json_str)
                logger.info(f"Successfully parsed EC2 specs: {ec2_specs}")
                
                # Ensure all instance types are in the result
                for it in instance_types:
                    if it not in ec2_specs:
                        ec2_specs[it] = {"vCPUs": None, "MemoryGiB": None}
                return ec2_specs
            else:
                logger.warning("No JSON found in EC2 specs response")
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error for EC2 specs: {e}")
        except Exception as e:
            logger.error(f"Error extracting EC2 specs: {e}")
        
        return {it: {"vCPUs": None, "MemoryGiB": None} for it in instance_types}

    def extract_rds_specs(self, instance_types: List[str]) -> Dict:
        if not instance_types:
            return {}

        prompt = f"""You are an AWS RDS specifications expert. Provide EXACT official AWS RDS instance specifications.

RDS instance types to lookup: {', '.join(instance_types)}

RDS instances follow similar patterns to EC2:
- db.t3.medium = 2 vCPUs, 4 GiB
- db.t3.large = 2 vCPUs, 8 GiB
- db.m5.large = 2 vCPUs, 8 GiB
- db.m5.xlarge = 4 vCPUs, 16 GiB
- db.r5.large = 2 vCPUs, 16 GiB
- db.r5.xlarge = 4 vCPUs, 32 GiB

Return ONLY valid JSON (no markdown, no text):
{{
  "db.t3.medium": {{"vCPUs": 2, "MemoryGiB": 4}},
  "db.m5.large": {{"vCPUs": 2, "MemoryGiB": 8}}
}}

JSON:"""

        try:
            response = call_groq(prompt, max_tokens=600)
            logger.info(f"RDS specs raw response: {response[:200]}")
            
            json_match = re.search(r'\{[^{}]*(?:\{[^{}]*\}[^{}]*)*\}', response, re.DOTALL)
            if json_match:
                json_str = json_match.group()
                rds_specs = json.loads(json_str)
                logger.info(f"Successfully parsed RDS specs: {rds_specs}")
                
                for it in instance_types:
                    if it not in rds_specs:
                        rds_specs[it] = {"vCPUs": None, "MemoryGiB": None}
                return rds_specs
            else:
                logger.warning("No JSON found in RDS specs response")
        except json.JSONDecodeError as e:
            logger.error(f"JSON decode error for RDS specs: {e}")
        except Exception as e:
            logger.error(f"Error extracting RDS specs: {e}")
        
        return {it: {"vCPUs": None, "MemoryGiB": None} for it in instance_types}

    def generate_service_description(self, service_name: str) -> str:
        prompt = f"""You are an AWS cloud expert. Describe this AWS service in ONE clear sentence.

Service: {service_name}

Format: {service_name}: [your description]

Be specific about what the service does and its primary use case."""

        try:
            response = call_groq(prompt, max_tokens=100)
            if response:
                # Clean up the response
                response = response.strip()
                # If response doesn't start with service name, add it
                if not response.startswith(service_name):
                    response = f"{service_name}: {response}"
                logger.info(f"Generated description for {service_name}")
                return response
            else:
                logger.warning(f"Empty response for {service_name}, using default")
                return f"{service_name}: Provides core cloud functionality."
        except Exception as e:
            logger.error(f"Error generating description for {service_name}: {e}")
            return f"{service_name}: Provides core cloud functionality."

    def generate_best_practices(self, services: List[str]) -> List[str]:
        if not services:
            return ["No specific services detected. General AWS best practices apply."]

        services_str = ", ".join(services[:10])  # Limit to first 10 services
        prompt = f"""You are an AWS Solutions Architect. Based on these AWS services: {services_str}

Provide 5 specific, actionable best practice recommendations focusing on:
- Cost optimization
- Security
- Performance
- Operational excellence

Format: Number each recommendation 1-5. Keep each to 1-2 sentences. Be specific to the services mentioned.

Start with "1." immediately:"""

        try:
            response = call_groq(prompt, max_tokens=500)
            if response:
                logger.info(f"Best practices response: {response[:100]}")
                # Extract numbered lines
                lines = []
                for line in response.split('\n'):
                    line = line.strip()
                    if line and re.match(r'^\d+\.', line):
                        lines.append(line)
                
                if lines:
                    logger.info(f"Found {len(lines)} best practice lines")
                    return lines[:5]
                else:
                    logger.warning("No numbered lines found in response")
            else:
                logger.warning("Empty response for best practices")
        except Exception as e:
            logger.error(f"Error generating best practices: {e}")
        
        # Fallback defaults
        return [
            "1. Implement IAM roles with least privilege principles for enhanced security.",
            "2. Enable CloudTrail and CloudWatch for comprehensive auditing and monitoring.",
            "3. Use AWS Cost Explorer and set up billing alerts for cost optimization.",
            "4. Implement auto-scaling and right-sizing for compute resources.",
            "5. Enable encryption at rest and in transit for all sensitive data."
        ]

    def extract_ec2_values(self, configuration_summary: str) -> tuple:
        try:
            os_match = re.search(r"operating\s*system\s*\((.*?)\)", configuration_summary, re.I)
            type_match = re.search(r"ec2\s*instance\s*\((.*?)\)", configuration_summary, re.I)
            price_match = re.search(r"pricing\s*strategy\s*\((.*?)\)", configuration_summary, re.I)
            return (
                os_match.group(1).strip() if os_match else None,
                type_match.group(1).strip() if type_match else None,
                price_match.group(1).strip() if price_match else None,
            )
        except:
            return None, None, None

    def extract_rds_values(self, configuration_summary: str, service_name: str) -> tuple:
        try:
            # Extract database type from service name (e.g., "Amazon RDS for MySQL" -> "MySQL")
            db_type = None
            db_engines = ["MySQL", "PostgreSQL", "MariaDB", "Oracle", "SQL Server", "Aurora"]
            for engine in db_engines:
                if engine.upper() in service_name.upper():
                    db_type = engine
                    break
            
            # Try multiple patterns to find instance type
            type_match = re.search(r"(?:rds\s*instance|instance\s*type|instance)\s*\((.*?)\)", configuration_summary, re.I)
            if not type_match:
                # Try to find db.* pattern anywhere in the config
                type_match = re.search(r"(db\.[a-z0-9]+\.[a-z0-9]+)", configuration_summary, re.I)
            
            price_match = re.search(r"(?:pricing\s*strategy|reserved|upfront)\s*\((.*?)\)", configuration_summary, re.I)
            if not price_match:
                # Try to extract pricing info from text
                if "reserved" in configuration_summary.lower():
                    if "no upfront" in configuration_summary.lower():
                        price_match = type('obj', (object,), {'group': lambda x: "Reserved No Upfront"})()
                    else:
                        price_match = type('obj', (object,), {'group': lambda x: "Reserved"})()
            
            return (
                db_type,
                type_match.group(1).strip() if type_match else None,
                price_match.group(1).strip() if price_match else None,
            )
        except:
            return None, None, None

    def generate_cost_report(self, input_file: str, output_file: str, customer_name: str,
                            usd_to_inr: float, region: str, pricing_link: str = ""):
        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        pricing_link_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        best_practice_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        note_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        thin = Side(style='thin')
        full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        try:
            if not os.path.exists(input_file):
                raise FileNotFoundError(f"Input CSV not found: {input_file}")

            df = pd.read_csv(input_file, skiprows=7)
            df.columns = df.columns.str.lower().str.strip()

            service_col = next((c for c in df.columns if "service" in c), None)
            monthly_col = next((c for c in df.columns if "monthly" in c or "cost" in c), None)
            config_col = next((c for c in df.columns if "configuration" in c or "summary" in c or "config" in c), None)

            if not all([service_col, monthly_col, config_col]):
                raise ValueError("Required columns not found")

            data = df[[service_col, monthly_col, config_col]].fillna("")

            has_ec2 = any("EC2" in str(row[service_col]).upper() for _, row in data.iterrows())
            has_rds = any("RDS" in str(row[service_col]).upper() for _, row in data.iterrows())

            instance_types = set()
            ec2_specs = {}
            if has_ec2:
                for val in data[config_col]:
                    if "EC2" in str(val).upper():
                        m = re.search(r"ec2\s*instance\s*\((.*?)\)", str(val), re.I)
                        if m:
                            instance_types.add(m.group(1).strip())
                ec2_specs = self.extract_ec2_specs(list(instance_types))

            rds_instance_types = set()
            rds_specs = {}
            if has_rds:
                for idx, row in data.iterrows():
                    if "RDS" in str(row[service_col]).upper():
                        config_val = str(row[config_col])
                        # Try multiple patterns
                        m = re.search(r"(?:rds\s*instance|instance\s*type|instance)\s*\((.*?)\)", config_val, re.I)
                        if not m:
                            m = re.search(r"(db\.[a-z0-9]+\.[a-z0-9]+)", config_val, re.I)
                        if m:
                            rds_instance_types.add(m.group(1).strip())
                if rds_instance_types:
                    rds_specs = self.extract_rds_specs(list(rds_instance_types))

            specs_failed = any(v["vCPUs"] is None and v["MemoryGiB"] is None for v in ec2_specs.values())
            rds_specs_failed = any(v["vCPUs"] is None and v["MemoryGiB"] is None for v in rds_specs.values())

            seen_services = set()
            cleaned_services = []
            for _, r in data.iterrows():
                svc = str(r[service_col]).strip()
                if svc and svc not in seen_services:
                    seen_services.add(svc)
                    clean_name = svc.split('(')[0].strip()
                    cleaned_services.append(clean_name)

            wb = openpyxl.Workbook()
            summary = wb.active
            summary.title = "Summary"

            summary.merge_cells("A1:B1")
            summary["A1"] = "Cost Estimation Summary"
            summary["A1"].font = Font(bold=True, color="000000")
            summary["A1"].alignment = Alignment(horizontal="center")
            summary["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            summary["A2"] = "Description"
            summary["B2"] = "Monthly Cost"
            for cell in [summary["A2"], summary["B2"]]:
                cell.fill = PatternFill(start_color="A7C5EB", end_color="A7C5EB", fill_type="solid")
                cell.border = full_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            total_usd = data[monthly_col].apply(pd.to_numeric, errors='coerce').sum()
            total_inr = total_usd * usd_to_inr

            summary["A3"] = "AWS Resource Cost (without TAX)"
            summary["B3"] = total_inr
            summary["B3"].number_format = "₹#,##0.00"
            summary["B3"].alignment = Alignment(horizontal="center")

            for r in [3]:
                for c in [1, 2]:
                    summary.cell(r, c).border = full_border

            summary.column_dimensions["A"].width = 30
            summary.column_dimensions["B"].width = 15

            sheet = wb.create_sheet("AWS Services")
            sheet.merge_cells("A1:K1" if (has_ec2 or has_rds) else "A1:E1")
            sheet["A1"] = f"Cost Estimation Report For {customer_name}"
            sheet["A1"].font = Font(bold=True, color="000000")
            sheet["A1"].alignment = Alignment(horizontal="center")
            sheet["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            if has_ec2 or has_rds:
                headers = [
                    "S.NO", "Instance Type", "vCPU", "RAM", "Operating System/Database",
                    "Running Hours", "Pricing Model", "Services", "Per Month USD", "Per Month INR", "Per Year INR"
                ]
                usd_col = 9
                inr_col = 10
                yearly_col = 11
            else:
                headers = ["S.NO", "Services", "Per Month USD", "Per Month INR", "Per Year INR"]
                usd_col = 3
                inr_col = 4
                yearly_col = 5

            for col, hdr in enumerate(headers, 1):
                cell = sheet.cell(2, col, hdr)
                cell.fill = header_fill
                cell.border = full_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            row = 3
            counter = 1
            rate_str = f"{usd_to_inr:.4f}"

            for _, r in data.iterrows():
                full_service = str(r[service_col]).strip()
                if not full_service:
                    continue
                try:
                    usd = float(r[monthly_col])
                except:
                    continue

                is_ec2 = "EC2" in full_service.upper()
                is_rds = "RDS" in full_service.upper()

                if (has_ec2 or has_rds) and (is_ec2 or is_rds):
                    if is_ec2:
                        os_val, instance_type, price_model = self.extract_ec2_values(r[config_col])
                        spec = ec2_specs.get(instance_type or "", {"vCPUs": None, "MemoryGiB": None})
                    else:  # is_rds
                        os_val, instance_type, price_model = self.extract_rds_values(r[config_col], full_service)
                        spec = rds_specs.get(instance_type or "", {"vCPUs": None, "MemoryGiB": None})

                    values = [
                        (1, counter), (2, instance_type or ""), (3, spec.get('vCPUs')),
                        (4, spec.get('MemoryGiB')), (5, os_val or ""), (6, "730 hours"),
                        (7, price_model or ""), (8, full_service), (9, usd)
                    ]

                    for col, val in values:
                        cell = sheet.cell(row, col, val)
                        cell.border = full_border
                        if col == 9:
                            cell.number_format = '$#,##0.00'
                        align_h = "right" if col in [2, 8] else ("left" if col < 9 else "right")
                        cell.alignment = Alignment(horizontal=align_h, vertical="center")

                    for col in [10, 11]:
                        formula = f"=I{row}*{rate_str}" if col == 10 else f"=J{row}*12"
                        cell = sheet.cell(row, col, value=formula)
                        cell.number_format = "₹#,##0.00"
                        cell.border = full_border
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                else:
                    sheet.cell(row, 1, counter).border = full_border
                    sheet.cell(row, 1).alignment = Alignment(horizontal="right", vertical="center")

                    merge_end = usd_col - 1
                    sheet.merge_cells(f"B{row}:{openpyxl.utils.get_column_letter(merge_end)}{row}")
                    cell = sheet.cell(row, 2, full_service)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    cell.border = full_border

                    for c in range(2, usd_col):
                        sheet.cell(row, c).border = full_border

                    sheet.cell(row, usd_col, usd).number_format = '$#,##0.00'
                    sheet.cell(row, usd_col).border = full_border
                    sheet.cell(row, usd_col).alignment = Alignment(horizontal="right", vertical="center")

                    sheet.cell(row, inr_col, value=f"={openpyxl.utils.get_column_letter(usd_col)}{row}*{rate_str}").number_format = "₹#,##0.00"
                    sheet.cell(row, inr_col).border = full_border
                    sheet.cell(row, inr_col).alignment = Alignment(horizontal="right", vertical="center")

                    sheet.cell(row, yearly_col, value=f"={openpyxl.utils.get_column_letter(inr_col)}{row}*12").number_format = "₹#,##0.00"
                    sheet.cell(row, yearly_col).border = full_border
                    sheet.cell(row, yearly_col).alignment = Alignment(horizontal="right", vertical="center")

                counter += 1
                row += 1

            last_data_row = row - 1

            total_r = row
            merge_end_total = "H" if (has_ec2 or has_rds) else openpyxl.utils.get_column_letter(usd_col - 1)
            sheet.merge_cells(f"A{total_r}:{merge_end_total}{total_r}")
            sheet.cell(total_r, 1, "Total Cost").alignment = Alignment(horizontal="right")
            sheet.cell(total_r, 1).border = full_border

            for c in range(1, (12 if (has_ec2 or has_rds) else usd_col + 3)):
                sheet.cell(total_r, c).border = full_border
                sheet.cell(total_r, c).fill = total_fill

            if last_data_row >= 3:
                sheet.cell(total_r, usd_col).value = f"=SUM({openpyxl.utils.get_column_letter(usd_col)}3:{openpyxl.utils.get_column_letter(usd_col)}{last_data_row})"
                sheet.cell(total_r, inr_col).value = f"=SUM({openpyxl.utils.get_column_letter(inr_col)}3:{openpyxl.utils.get_column_letter(inr_col)}{last_data_row})"
                sheet.cell(total_r, yearly_col).value = f"=SUM({openpyxl.utils.get_column_letter(yearly_col)}3:{openpyxl.utils.get_column_letter(yearly_col)}{last_data_row})"
            else:
                sheet.cell(total_r, usd_col).value = 0
                sheet.cell(total_r, inr_col).value = 0
                sheet.cell(total_r, yearly_col).value = 0

            sheet.cell(total_r, usd_col).number_format = '$#,##0.00'
            sheet.cell(total_r, inr_col).number_format = '₹#,##0.00'
            sheet.cell(total_r, yearly_col).number_format = '₹#,##0.00'

            pl_row = total_r + 1
            merge_end_pl = "H" if (has_ec2 or has_rds) else openpyxl.utils.get_column_letter(usd_col - 1)
            sheet.merge_cells(f"A{pl_row}:{merge_end_pl}{pl_row}")
            sheet.cell(pl_row, 1, "Pricing Link").alignment = Alignment(horizontal="right")

            for c in range(1, (12 if (has_ec2 or has_rds) else usd_col + 3)):
                sheet.cell(pl_row, c).border = full_border
                sheet.cell(pl_row, c).fill = pricing_link_fill

            sheet.cell(pl_row, usd_col if not (has_ec2 or has_rds) else 9, pricing_link or "Not provided")

            note_row = pl_row + 1
            note_title = sheet.cell(note_row, 1, "Note:")
            note_title.fill = note_fill

            notes = [
                "1. The given costs are considered as estimation based on the Monthly usage actual amount will get vary.",
                f"2. {region} region is considered for this workload.",
                f"3. The exchange rate is considered as ₹{usd_to_inr:.2f} as per {datetime.now().strftime('%d/%m/%y')} date."
            ]

            for i, note_text in enumerate(notes, 1):
                sheet.cell(note_row + i, 1, note_text)

            cur_note_row = note_row + len(notes) + 1
            note_sno = 4

            if specs_failed and has_ec2:
                sheet.cell(cur_note_row, 1, f"{note_sno}. Failed to extract EC2 specs (vCPU, RAM) from model.")
                cur_note_row += 1
                note_sno += 1

            if rds_specs_failed and has_rds:
                sheet.cell(cur_note_row, 1, f"{note_sno}. Failed to extract RDS specs (vCPU, RAM) from model.")
                cur_note_row += 1
                note_sno += 1

            seen = set()
            for _, r in data.iterrows():
                svc = str(r[service_col]).strip()
                if svc and svc not in seen:
                    seen.add(svc)
                    clean = svc.split('(')[0].strip()
                    desc = self.generate_service_description(clean)
                    desc = desc.replace(f"{clean}:", "").strip()
                    sheet.cell(cur_note_row, 1, f"{note_sno}. {clean}: {desc}")
                    cur_note_row += 1
                    note_sno += 1

            bp_row = cur_note_row
            bp_title = sheet.cell(bp_row, 1, "Best Practices:")
            bp_title.fill = best_practice_fill

            best_practices = self.generate_best_practices(cleaned_services)

            for i, line in enumerate(best_practices, 1):
                sheet.cell(bp_row + i, 1, line)

            if has_ec2 or has_rds:
                widths = [6, 28, 10, 10, 18, 14, 16, 40, 14, 14, 14]
            else:
                widths = [6, 50, 16, 16, 16]

            for col, width in enumerate(widths, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

            wb.save(output_file)
            logger.info(f"Report saved: {output_file}")
            return {"status": "success", "file": output_file}

        except Exception as e:
            logger.exception("Report generation failed")
            return {"status": "error", "message": str(e)}
