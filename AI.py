import streamlit as st
import boto3
import pandas as pd
import io
import json
import logging
from botocore.exceptions import ClientError
from functools import lru_cache
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
from datetime import datetime
import os
import tempfile
import uuid

# Set page configuration for the app
st.set_page_config(page_title="GenAI First Implementation", page_icon=":rocket:", layout="wide")

# Custom CSS for professional and attractive design, including logo positioning
st.markdown("""
    <style>
    /* Main container styling */
    .main {
        background-color: #f5f6f5;
        padding: 40px;
        border-radius: 15px;
        box-shadow: 0 4px 8px rgba(0, 0, 0, 0.1);
        min-height: 100vh;
        display: flex;
        flex-direction: column;
        align-items: center;
        justify-content: center;
    }
    /* Heading styling */
    .stTitle {
        color: #232f3e;
        font-size: 48px;
        font-weight: bold;
        text-align: center;
        margin-bottom: 40px;
        text-shadow: 1px 1px 2px rgba(0, 0, 0, 0.1);
    }
    /* Button container */
    .button-container {
        display: flex;
        gap: 20px;
        justify-content: center;
    }
    /* Button styling */
    .stButton>button {
        background-color: #ff9900;
        color: #ffffff;
        font-size: 18px;
        font-weight: bold;
        padding: 15px 30px;
        border-radius: 10px;
        border: none;
        transition: background-color 0.3s ease, transform 0.2s ease;
    }
    .stButton>button:hover {
        background-color: #e68a00;
        transform: scale(1.05);
    }
    /* Subtle background gradient */
    body {
        background: linear-gradient(135deg, #FFDAB9 0%, #D3D3D3 100%);
    }
    /* Form styling for apps */
    .stForm {
        background-color: #FFFFFF;
        border: 2px solid #D9D9D9;
        border-radius: 10px;
        padding: 20px;
    }
    /* Success message styling */
    .stSuccess {
        background-color: #FFFFFF;
        color: #3C763D;
        border-radius: 5px;
        padding: 10px;
    }
    /* Error message styling */
    стеррор {
        background-color: #FFFFFF;
        color: #A94442;
        border-radius: 5px;
        padding: 10px;
    }
    /* Dataframe styling */
    .stDataFrame {
        border: 1px solid #D9D9D9;
        border-radius: 5px;
    }
    /* Logo styling */
    .logo-container {
        position: absolute;
        top: 20px;
        right: 20px;
    }
    .logo-img {
        max-width: 150px;
        height: auto;
    }
    </style>
""", unsafe_allow_html=True)

# Initialize session state for page navigation
if 'page' not in st.session_state:
    st.session_state.page = 'landing'

# Landing Page
def landing_page():
    # Add logo in the top-right corner
    logo_path = "C:/Users/KarthigaS/Downloads/GenAI-first/logo.php"  # Adjust path as needed
    if os.path.exists(logo_path):
        st.markdown('<div class="logo-container">', unsafe_allow_html=True)
        st.image(logo_path, width=150, use_container_width=False, clamp=True, output_format="auto")
        st.markdown('</div>', unsafe_allow_html=True)
    else:
        st.warning("Logo file not found. Please ensure the file exists at the specified path.")

    st.title("GenAI First Implementation")
    
    st.markdown("""
        <div class='button-container'>
    """, unsafe_allow_html=True)
    
    col1, col2 = st.columns(2)
    
    with col1:
        if st.button("Modernization Pathway"):
            st.session_state.page = 'modernization'
    
    with col2:
        if st.button("Pricing Calculator"):
            st.session_state.page = 'pricing'
    
    st.markdown("</div>", unsafe_allow_html=True)

# Modernization Pathway App
def modernization_pathway():
    # Configure logging
    logging.basicConfig(level=logging.INFO)
    logger = logging.getLogger(__name__)
    
    # AWS Configuration
    AWS_REGION = "us-east-1"
    S3_BUCKET_NAME = "modernization-assessment"
    S3_FILE_KEY = "Security_Questions_App_Infra.xlsx"
    BEDROCK_MODEL_ID = "amazon.titan-text-premier-v1:0"
    BLOG_LINK = "https://aws.amazon.com/blogs/migration-and-modernization/move-to-ai-pathway/"

    def download_excel_file_from_s3():
        """Downloads the Excel file from the specified S3 bucket."""
        try:
            s3_client = boto3.client("s3", region_name=AWS_REGION)
            obj = s3_client.get_object(Bucket=S3_BUCKET_NAME, Key=S3_FILE_KEY)
            file_content = io.BytesIO(obj["Body"].read())
            return pd.ExcelFile(file_content)
        except Exception as e:
            st.error(f"Error accessing S3 file: {e}")
            logger.error(f"S3 error: {e}")
            return None

    @lru_cache(maxsize=32)
    def fetch_questions_from_excel(excel_file):
        """Fetches questions from the specified Excel file."""
        modules = {}
        try:
            sheet_name = excel_file.sheet_names[0]
            df = excel_file.parse(sheet_name, header=None)
            questions_col = df.iloc[:, 1].dropna().tolist()
            current_module = None
            for question in questions_col:
                if any(keyword in question for keyword in ["Tech. Stack", "Security", "Code Quality", "Testing", "AWS Infra."]):
                    current_module = question.split(':')[0].strip()
                    modules[current_module] = []
                elif current_module:
                    modules[current_module].append(question)
        except Exception as e:
            st.error(f"Error reading Excel file: {e}")
            logger.error(f"Excel parsing error: {e}")
            return {}
        return modules

    def invoke_bedrock_model(input_text):
        """Invokes Amazon Bedrock Titan Text Premier model to generate recommendations."""
        try:
            bedrock_client = boto3.client("bedrock-runtime", region_name=AWS_REGION)
            payload = {
                "modelId": BEDROCK_MODEL_ID,
                "contentType": "application/json",
                "accept": "application/json",
                "body": json.dumps({
                    "inputText": input_text,
                    "textGenerationConfig": {
                        "maxTokenCount": 3072,
                        "stopSequences": [],
                        "temperature": 0.7,
                        "topP": 0.9
                    }
                })
            }
            logger.info(f"Invoking Bedrock model with payload: {payload}")
            response = bedrock_client.invoke_model(**payload)
            response_body = json.loads(response["body"].read())
            logger.info(f"Bedrock response: {response_body}")
            if "results" in response_body and len(response_body["results"]) > 0:
                return response_body["results"][0].get("outputText", "No response from model.")
            else:
                logger.warning("No results found in Bedrock response.")
                st.error("Failed to get response from Bedrock model.")
                return "Unable to generate recommendation at this time."
        except ClientError as e:
            st.error(f"Error invoking Bedrock model: {e}")
            logger.error(f"Bedrock error: {e}")
            return "Unable to generate recommendation due to an error."

    def generate_recommendations_batch(user_responses):
        """Generates recommendations for all questions in a single model invocation."""
        combined_input = " ".join(user_responses.values())
        prompt = f"""Analyze the following responses and generate detailed recommendations in the following format:
            1. AWS Service Recommendation: 
                - Compute
                - Database
                - Storage
                - CI/CD
                - Integration Services
            2. If already using AWS services, give best practices
            3. Security best practices
            4. Cost Optimization techniques
            Responses:
            {combined_input}
        """
        return invoke_bedrock_model(prompt)

    def display_recommendations(summary, modules_with_questions, user_responses):
        """Displays a structured summary of recommendations."""
        st.subheader("Quadra Modernization Insights and Advice")
        if summary:
            st.markdown(summary)
        else:
            for module, questions in modules_with_questions.items():
                st.markdown(f"### {module} Recommendations")
                for question in questions:
                    st.markdown(f"**Question: {question} - Your Response: {user_responses[question]}**")
                    st.write("Generating recommendation...")  # Placeholder text

    def answer_additional_question(question):
        """Generate an answer using an intelligent AI service."""
        try:
            response = invoke_bedrock_model(f"Answer this question: {question}")
            return response
        except Exception as e:
            logger.error(f"Error answering question: {e}")
            return "Unable to generate a response at this time."

    def main_modernization():
        st.title("AWS Modernization Assessment Tool")
        st.header("Current Application Assessment Questions")
        excel_file = download_excel_file_from_s3()
        if not excel_file:
            return
        modules_with_questions = fetch_questions_from_excel(excel_file)
        if not modules_with_questions:
            st.error("No questions found in the Excel file.")
            return
        user_responses = {}
        for module, questions in modules_with_questions.items():
            st.markdown(f"### {module} Questions")
            for question in questions:
                response = st.text_area(f"{question}", key=question)
                user_responses[question] = response if response else "No response provided"
        all_filled = all(value.strip() != "" for value in user_responses.values())
        if all_filled:
            if st.button("Submit Responses"):
                st.success("Responses submitted successfully!")
                with st.spinner('Generating recommendations...'):
                    summary = generate_recommendations_batch(user_responses)
                display_recommendations(summary, modules_with_questions, user_responses)
            
            # Additional questions section
            st.markdown("#### Do you have any additional questions?")
            additional_question = st.text_input("Enter your question here:")
            if additional_question:
                answer = answer_additional_question(additional_question)
                st.markdown(f"**Answer:** {answer}")
        else:
            st.info("Please fill in all responses to enable submission.")
        
        # Back button to return to landing page
        if st.button("Back to Home"):
            st.session_state.page = 'landing'

    main_modernization()

# Pricing Calculator App
def pricing_calculator():
    # Set up logging
    logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
    logger = logging.getLogger(__name__)

    # Initialize AWS clients
    bedrock_client = boto3.client("bedrock-runtime", region_name="us-east-1")
    LITE_MODEL_ID = "us.amazon.nova-lite-v1:0"

    class CostReportAgent:
        def __init__(self, default_usd_to_inr: float, default_region: str = "US East (N. Virginia)"):
            self.usd_to_inr = default_usd_to_inr
            self.default_region = default_region
            self.default_output_dir = tempfile.gettempdir()
            self.tools = {
                "extract_ec2_specs": self.extract_ec2_specs,
                "generate_service_description": self.generate_service_description,
                "generate_cost_report": self.generate_cost_report,
            }
            logger.info("CostReportAgent initialized with USD to INR rate: %.2f", self.usd_to_inr)

        def extract_ec2_specs(self, instance_types: list[str]) -> dict:
            """Extract EC2 instance specs using Bedrock Nova Lite model."""
            if not instance_types:
                logger.warning("No EC2 instance types provided. Returning empty specs.")
                return {}

            system_prompt = [
                {
                    "text": """
                    You are a data extraction assistant specialized in AWS EC2 instances. Given a list of EC2 instance types, provide the vCPUs and Memory (GiB) for each instance type. Return the result as a JSON object where keys are the instance types and values are dictionaries with 'vCPUs' and 'MemoryGiB' keys. Use null if a value is not found. Ensure all specified instance types are included in the output, even if their values are not found.
                    """
                }
            ]
            user_message = [
                {"role": "user", "content": [{"text": f"Instance Types: {', '.join(instance_types)}"}]}
            ]
            inference_params = {"maxTokens": 500, "topP": 0.9, "topK": 20, "temperature": 0.7}
            request_body = {
                "schemaVersion": "messages-v1",
                "messages": user_message,
                "system": system_prompt,
                "inferenceConfig": inference_params,
            }

            try:
                response = bedrock_client.invoke_model_with_response_stream(
                    modelId=LITE_MODEL_ID, body=json.dumps(request_body)
                )
                full_response = ""
                for event in response.get("body", []):
                    chunk = event.get("chunk")
                    if chunk:
                        chunk_json = json.loads(chunk.get("bytes").decode())
                        if content_block_delta := chunk_json.get("contentBlockDelta"):
                            if text := content_block_delta.get("delta", {}).get("text"):
                                full_response += text
                logger.debug("Bedrock response: %s", full_response)

                json_start = full_response.find("```json") + 7
                json_end = full_response.rfind("```")
                if json_start > 6 and json_end > json_start:
                    ec2_specs = json.loads(full_response[json_start:json_end].strip())
                    for instance_type in instance_types:
                        if instance_type not in ec2_specs:
                            ec2_specs[instance_type] = {"vCPUs": None, "MemoryGiB": None}
                else:
                    logger.warning("No JSON block found in Bedrock response.")
                    ec2_specs = {it: {"vCPUs": None, "MemoryGiB": None} for it in instance_types}
            except (ClientError, json.JSONDecodeError) as e:
                logger.error("Error extracting EC2 specs: %s", e)
                ec2_specs = {it: {"vCPUs": None, "MemoryGiB": None} for it in instance_types}

            logger.info("Extracted EC2 specs: %s", ec2_specs)
            return ec2_specs

        def generate_service_description(self, service_name: str) -> str:
            """Generate a concise description for an AWS service using Nova Lite model."""
            system_prompt = [
                {
                    "text": """
                    You are an AI assistant specialized in AWS services. Given the name of an AWS service, provide a highly precise, concise description (one sentence) of its primary purpose in a cloud environment. Start with the service name, followed by a colon and the reason, using exact functionality without vague terms or extra details. Generate a unique description based on the service name provided, without relying on predefined examples.
                    """
                }
            ]
            user_message = [{"role": "user", "content": [{"text": f"Service: {service_name}"}]}]
            inference_params = {"maxTokens": 50, "topP": 0.9, "topK": 20, "temperature": 0.5}
            request_body = {
                "schemaVersion": "messages-v1",
                "messages": user_message,
                "system": system_prompt,
                "inferenceConfig": inference_params,
            }

            try:
                response = bedrock_client.invoke_model_with_response_stream(
                    modelId=LITE_MODEL_ID, body=json.dumps(request_body)
                )
                full_response = ""
                for event in response.get("body", []):
                    chunk = event.get("chunk")
                    if chunk:
                        chunk_json = json.loads(chunk.get("bytes").decode())
                        if content_block_delta := chunk_json.get("contentBlockDelta"):
                            if text := content_block_delta.get("delta", {}).get("text"):
                                full_response += text
                logger.debug("Service description response for %s: %s", service_name, full_response)
                return full_response.strip()
            except ClientError as e:
                logger.error("Error generating description for %s: %s", service_name, e)
                return f"{service_name}: Provides core cloud functionality."

        def generate_best_practices(self, services: list[str]) -> list[str]:
            """Generate best practices for the given AWS services using Bedrock Nova Lite model."""
            if not services:
                logger.warning("No services provided for best practices generation. Returning default best practices.")
                return [
                    "1. Implement AWS Identity and Access Management (IAM) roles with least privilege principles to enhance security for all services.",
                    "2. Enable AWS CloudTrail to monitor and log API activity for auditing and compliance purposes.",
                    "3. Use AWS Cost Explorer to regularly review and optimize costs based on usage patterns and recommendations.",
                    "4. Configure Amazon CloudWatch for real-time monitoring and alerting to maintain service performance and security."
                ]

            system_prompt = [
                {
                    "text": """
                    You are an AWS architecture expert. Given a list of AWS services, generate a list of concise, specific best practices for optimizing security and infrastructure for those services. Each best practice should be one sentence, focus on security or infrastructure optimization, and be directly relevant to the provided services. Return the result as a JSON array of strings, with each string prefixed by a number (e.g., "1. ", "2. "). Limit to 5 best practices total, prioritizing the most critical security and infrastructure recommendations.
                    """
                }
            ]
            user_message = [
                {"role": "user", "content": [{"text": f"Services: {', '.join(services)}"}]}
            ]
            inference_params = {"maxTokens": 500, "topP": 0.9, "topK": 20, "temperature": 0.6}
            request_body = {
                "schemaVersion": "messages-v1",
                "messages": user_message,
                "system": system_prompt,
                "inferenceConfig": inference_params,
            }

            try:
                response = bedrock_client.invoke_model_with_response_stream(
                    modelId=LITE_MODEL_ID, body=json.dumps(request_body)
                )
                full_response = ""
                for event in response.get("body", []):
                    chunk = event.get("chunk")
                    if chunk:
                        chunk_json = json.loads(chunk.get("bytes").decode())
                        if content_block_delta := chunk_json.get("contentBlockDelta"):
                            if text := content_block_delta.get("delta", {}).get("text"):
                                full_response += text
                logger.debug("Best practices response: %s", full_response)

                json_start = full_response.find("```json") + 7
                json_end = full_response.rfind("```")
                if json_start > 6 and json_end > json_start:
                    best_practices = json.loads(full_response[json_start:json_end].strip())
                    if not isinstance(best_practices, list):
                        raise ValueError("Expected JSON array for best practices")
                    return best_practices[:5]  # Limit to 5 best practices
                else:
                    logger.warning("No JSON block found in Bedrock response for best practices.")
                    return [
                        "1. Implement AWS Identity and Access Management (IAM) roles with least privilege principles to enhance security for all services.",
                        "2. Enable AWS CloudTrail to monitor and log API activity for auditing and compliance purposes.",
                        "3. Use AWS Cost Explorer to regularly review and optimize costs based on usage patterns and recommendations.",
                        "4. Configure Amazon CloudWatch for real-time monitoring and alerting to maintain service performance and security."
                    ]
            except (ClientError, json.JSONDecodeError, ValueError) as e:
                logger.error("Error generating best practices: %s", e)
                return [
                    "1. Implement AWS Identity and Access Management (IAM) roles with least privilege principles to enhance security for all services.",
                    "2. Enable AWS CloudTrail to monitor and log API activity for auditing and compliance purposes.",
                    "3. Use AWS Cost Explorer to regularly review and optimize costs based on usage patterns and recommendations.",
                    "4. Configure Amazon CloudWatch for real-time monitoring and alerting to maintain service performance and security."
                ]

        def extract_ec2_values(self, configuration_summary: str) -> tuple:
            """Extract Operating System, EC2 Type, and Pricing Model from configuration summary."""
            try:
                os_value = re.search(r"operating\s*system\s*\((.*?)\)", configuration_summary, re.IGNORECASE)
                ec2_type = re.search(r"ec2\s*instance\s*\((.*?)\)", configuration_summary, re.IGNORECASE)
                pricing_model = re.search(r"pricing\s*strategy\s*\((.*?)\)", configuration_summary, re.IGNORECASE)
                return (
                    os_value.group(1) if os_value else None,
                    ec2_type.group(1) if ec2_type else None,
                    pricing_model.group(1) if pricing_model else None,
                )
            except AttributeError:
                return None, None, None

        def generate_cost_report(self, input_file: str, output_file: str, customer_name: str, usd_to_inr: float | None = None, region: str | None = None, pricing_link: str | None = None):
            """Generate an Excel cost report from a CSV input file after removing first 7 rows."""
            usd_to_inr = usd_to_inr or self.usd_to_inr
            region = region or self.default_region
            logger.info("Generating cost report: input=%s, output=%s, customer_name=%s, usd_to_inr=%.2f, region=%s, pricing_link=%s", 
                        input_file, output_file, customer_name, usd_to_inr, region, pricing_link)

            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_file), exist_ok=True)

            # Define styles for Excel
            header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
            total_gst_fill = PatternFill(start_color="A7C5EB", end_color="A7C5EB", fill_type="solid")
            pricing_link_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            border = Border(
                left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
            )
            merged_bottom_border = Border(bottom=Side(style="thin"))

            try:
                if not os.path.exists(input_file):
                    raise FileNotFoundError(f"Input file {input_file} does not exist.")

                # Read CSV and skip first 7 rows
                df = pd.read_csv(input_file, skiprows=7)
                df.columns = df.columns.str.lower()
                service_column = next((col for col in df.columns if "service" in col.lower()), None)
                monthly_column = next((col for col in df.columns if "monthly" in col.lower()), None)
                config_summary_column = next((col for col in df.columns if "configuration summary" in col.lower()), None)

                if not all([service_column, monthly_column, config_summary_column]):
                    raise ValueError("Required columns (service, monthly cost, configuration summary) not found in CSV.")

                services_data = df[[service_column, monthly_column, config_summary_column]].fillna("")
                instance_types = {
                    match.group(1)
                    for row in services_data[config_summary_column]
                    if pd.notna(row) and "EC2" in row
                    for match in [re.search(r"ec2\s*instance\s*\((.*?)\)", row, re.IGNORECASE)] if match
                }
                logger.info("Detected EC2 instance types: %s", instance_types)

                ec2_services = []
                other_services = []
                for i, row in services_data.iterrows():
                    service_name = row[service_column]
                    if pd.notna(service_name) and service_name.strip():
                        if "EC2" in service_name.upper():
                            if service_name not in [s[1] for s in ec2_services]:
                                ec2_services.append((i + 1, service_name))
                        else:
                            if service_name not in [s[1] for s in other_services]:
                                other_services.append((i + 1, service_name))
                service_list = ec2_services + other_services
                cleaned_service_list = [service_name.split("(")[0].strip() for _, service_name in service_list]
                logger.info("Service list: %s", service_list)

                ec2_specs = self.extract_ec2_specs(instance_types)
                specs_extraction_failed = any(
                    spec["vCPUs"] is None and spec["MemoryGiB"] is None for spec in ec2_specs.values()
                )

                workbook = openpyxl.Workbook()
                summary_sheet = workbook.active
                summary_sheet.title = "Summary"

                # Populate Summary sheet
                summary_sheet.merge_cells("A1:B1")
                summary_sheet["A1"] = "Cost Estimation Summary"
                summary_sheet["A1"].font = Font(bold=True, color="000000")
                summary_sheet["A1"].alignment = Alignment(horizontal="center")
                summary_sheet["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                summary_sheet["A2"] = "Description"
                summary_sheet["B2"] = "Monthly Cost"
                for col in range(1, 3):
                    cell = summary_sheet.cell(row=2, column=col)
                    cell.fill = PatternFill(start_color="A7C5EB", end_color="A7C5EB", fill_type="solid")
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")
                    cell.font = Font(bold=True, color="000000")

                total_cost_usd = services_data[monthly_column].apply(pd.to_numeric, errors="coerce").sum()
                total_cost_inr = total_cost_usd * usd_to_inr
                summary_sheet["A3"] = "AWS Resource Cost (without TAX)"
                cell = summary_sheet.cell(row=3, column=2, value=total_cost_inr)
                cell.number_format = "₹ #,##0.00"
                for col in range(1, 3):
                    cell = summary_sheet.cell(row=3, column=col)
                    cell.border = border
                    cell.alignment = Alignment(horizontal="center")

                summary_sheet.column_dimensions["A"].width = 30
                summary_sheet.column_dimensions["B"].width = 15

                aws_services_sheet = workbook.create_sheet("AWS Services")
                aws_services_sheet.merge_cells("A1:K1")
                aws_services_sheet["A1"] = f"Cost Estimation Report For {customer_name}"
                aws_services_sheet["A1"].font = Font(bold=True, color="000000")
                aws_services_sheet["A1"].alignment = Alignment(horizontal="center")
                aws_services_sheet["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

                headers = [
                    "S.NO", "EC2 Type (Advanced EC2 Instance)", "vCPU", "RAM", "Operating System",
                    "Running Hours", "Pricing Model", "Services", "Per Month USD", "Per Month INR", "Per Year INR"
                ]
                for col_num, header in enumerate(headers, 1):
                    cell = aws_services_sheet.cell(row=2, column=col_num, value=header)
                    cell.fill = header_fill
                    cell.border = border
                    cell.font = Font(bold=True)

                current_row = 3
                ec2_rows = []
                other_rows = []
                ec2_index = 1

                for i, row in services_data.iterrows():
                    service_name = row[service_column]
                    usd_value = row[monthly_column]
                    if pd.isna(usd_value) or usd_value == "":
                        continue

                    if "EC2" in service_name.upper():
                        usd_value = float(usd_value)
                        inr_value = usd_value * usd_to_inr
                        yearly_inr_value = inr_value * 12
                        os_value, ec2_type, pricing_model = self.extract_ec2_values(row[config_summary_column])
                        specs = ec2_specs.get(ec2_type, {"vCPUs": None, "MemoryGiB": None})
                        ec2_rows.append({
                            "row": current_row, "sno": ec2_index, "ec2_type": ec2_type, "vcpu": specs.get("vCPUs"),
                            "ram": specs.get("MemoryGiB"), "os": os_value, "running_hours": "730 hours",
                            "pricing_model": pricing_model, "service": service_name, "usd": usd_value,
                            "inr": inr_value, "yearly_inr": yearly_inr_value
                        })
                        ec2_index += 1
                        current_row += 1

                other_index = ec2_index
                for i, row in services_data.iterrows():
                    service_name = row[service_column]
                    usd_value = row[monthly_column]
                    if pd.isna(usd_value) or usd_value == "":
                        continue
                    if "EC2" not in service_name.upper():
                        usd_value = float(usd_value)
                        inr_value = usd_value * usd_to_inr
                        yearly_inr_value = inr_value * 12
                        other_rows.append({
                            "row": current_row, "sno": other_index, "service": service_name,
                            "usd": usd_value, "inr": inr_value, "yearly_inr": yearly_inr_value
                        })
                        other_index += 1
                        current_row += 1

                for ec2_row in ec2_rows:
                    row = ec2_row["row"]
                    cells = [
                        (1, ec2_row["sno"]), (2, ec2_row["ec2_type"] or ""), (3, ec2_row["vcpu"] or ""),
                        (4, ec2_row["ram"] or ""), (5, ec2_row["os"] or ""), (6, ec2_row["running_hours"] or ""),
                        (7, ec2_row["pricing_model"] or ""), (8, ec2_row["service"]), (9, f"$ {ec2_row['usd']:.2f}"),
                        (10, ec2_row["inr"]), (11, ec2_row["yearly_inr"])
                    ]
                    for col, value in cells:
                        cell = aws_services_sheet.cell(row=row, column=col, value=value)
                        cell.border = border
                        cell.alignment = Alignment(horizontal="left" if col <= 8 else "right", vertical="center")
                        if col >= 10:
                            cell.number_format = "₹ #,##0.00"

                for other_row in other_rows:
                    row = other_row["row"]
                    aws_services_sheet.cell(row=row, column=1, value=other_row["sno"]).border = border
                    aws_services_sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=8)
                    cell = aws_services_sheet.cell(row=row, column=2, value=other_row["service"])
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    for col in range(2, 9):
                        aws_services_sheet.cell(row=row, column=col).border = merged_bottom_border
                    aws_services_sheet.cell(row=row, column=9, value=f"$ {other_row['usd']:.2f}").border = border
                    cell = aws_services_sheet.cell(row=row, column=10, value=other_row["inr"])
                    cell.border = border
                    cell.number_format = "₹ #,##0.00"
                    cell = aws_services_sheet.cell(row=row, column=11, value=other_row["yearly_inr"])
                    cell.border = border
                    cell.number_format = "₹ #,##0.00"

                total_row = current_row
                total_cost_usd = services_data[monthly_column].apply(pd.to_numeric, errors="coerce").sum()
                total_cost_inr = total_cost_usd * usd_to_inr
                total_cost_yearly_inr = total_cost_inr * 12

                aws_services_sheet.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=8)
                aws_services_sheet.cell(row=total_row, column=1, value="Total Cost").alignment = Alignment(horizontal="right")
                for col in range(1, 12):
                    cell = aws_services_sheet.cell(row=total_row, column=col)
                    cell.border = border
                    cell.fill = total_fill
                aws_services_sheet.cell(row=total_row, column=9, value=f"$ {total_cost_usd:.2f}")
                cell = aws_services_sheet.cell(row=total_row, column=10, value=total_cost_inr)
                cell.number_format = "₹ #,##0.00"
                cell = aws_services_sheet.cell(row=total_row, column=11, value=total_cost_yearly_inr)
                cell.number_format = "₹ #,##0.00"

                total_cost_gst_usd = total_cost_usd * 1.18
                total_cost_gst_inr = total_cost_inr * 1.18
                total_cost_gst_yearly_inr = total_cost_yearly_inr * 1.18
                aws_services_sheet.merge_cells(start_row=total_row + 1, start_column=1, end_row=total_row + 1, end_column=8)
                aws_services_sheet.cell(row=total_row + 1, column=1, value="Total Cost with GST").alignment = Alignment(horizontal="right")
                for col in range(1, 12):
                    cell = aws_services_sheet.cell(row=total_row + 1, column=col)
                    cell.border = border
                    cell.fill = total_gst_fill
                aws_services_sheet.cell(row=total_row + 1, column=9, value=f"$ {total_cost_gst_usd:.2f}")
                cell = aws_services_sheet.cell(row=total_row + 1, column=10, value=total_cost_gst_inr)
                cell.number_format = "₹ #,##0.00"
                cell = aws_services_sheet.cell(row=total_row + 1, column=11, value=total_cost_gst_yearly_inr)
                cell.number_format = "₹ #,##0.00"

                # Add Pricing Link row
                pricing_link_row = total_row + 2
                aws_services_sheet.merge_cells(start_row=pricing_link_row, start_column=1, end_row=pricing_link_row, end_column=8)
                aws_services_sheet.cell(row=pricing_link_row, column=1, value="Pricing Link").alignment = Alignment(horizontal="right")
                for col in range(1, 12):
                    cell = aws_services_sheet.cell(row=pricing_link_row, column=col)
                    cell.border = border
                    cell.fill = pricing_link_fill
                aws_services_sheet.cell(row=pricing_link_row, column=9, value=pricing_link or "Not provided")

                note_row = pricing_link_row + 1
                aws_services_sheet.cell(row=note_row, column=1, value="Note:").fill = PatternFill(
                    start_color="FFFF00", end_color="FFFF00", fill_type="solid"
                )
                note1 = "1. The given costs are considered as estimation based on the Monthly usage actual amount will get vary."
                note2 = f"2. {region} region is considered for this workload."
                note3 = f"3. The exchange rate is considered as ₹{usd_to_inr:.2f} as per {datetime.now().strftime('%d/%m/%y')} date."
                aws_services_sheet.cell(row=note_row + 1, column=1, value=note1)
                aws_services_sheet.cell(row=note_row + 2, column=1, value=note2)
                aws_services_sheet.cell(row=note_row + 3, column=1, value=note3)

                current_note_row = note_row + 4
                note_sno = 4
                if specs_extraction_failed:
                    failure_note = f"{note_sno}. Failed to extract EC2 specs (vCPU, RAM) from Nova Lite model."
                    aws_services_sheet.cell(row=current_note_row, column=1, value=failure_note)
                    current_note_row += 1
                    note_sno += 1

                for sno, service_name in service_list:
                    cleaned_service_name = service_name.split("(")[0].strip()
                    description = self.generate_service_description(cleaned_service_name)
                    description_cleaned = description.replace(f"{cleaned_service_name}:", "").strip()
                    service_note = f"{note_sno}. {cleaned_service_name}: {description_cleaned}"
                    aws_services_sheet.cell(row=current_note_row, column=1, value=service_note)
                    current_note_row += 1
                    note_sno += 1

                # Adding Best Practices Section
                best_practice_row = current_note_row
                aws_services_sheet.cell(row=best_practice_row, column=1, value="Best Practices:").fill = PatternFill(
                    start_color="90EE90", end_color="90EE90", fill_type="solid"
                )
                best_practice_notes = self.generate_best_practices(cleaned_service_list)
                for idx, note in enumerate(best_practice_notes, start=1):
                    aws_services_sheet.cell(row=best_practice_row + idx, column=1, value=note)

                for col in aws_services_sheet.columns:
                    max_length = 0
                    col_letter = openpyxl.utils.get_column_letter(col[0].column)
                    if col_letter == "A":
                        aws_services_sheet.column_dimensions[col_letter].width = 6
                        continue
                    elif col_letter == "B":
                        aws_services_sheet.column_dimensions[col_letter].width = 20
                        continue
                    for cell in col:
                        if not isinstance(cell, openpyxl.cell.cell.MergedCell) and cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                    aws_services_sheet.column_dimensions[col_letter].width = max_length + 2

                workbook.save(output_file)
                logger.info("Cost report generated successfully at %s", output_file)
                return {"status": "success", "output_file": output_file}

            except Exception as e:
                logger.error("Error generating cost report: %s", e)
                return {"status": "error", "message": str(e)}

    def main_pricing():
        st.title("AWS Pricing Calculator")
        st.markdown("""
        Upload a CSV file containing AWS service cost data to generate a detailed cost report in Excel format.
        The report includes EC2 instance details, cost breakdowns in USD and INR, service descriptions, a pricing link, and best practices for cost optimization and security.
        Ensure the CSV contains columns for 'Service', 'Monthly Cost', and 'Configuration Summary'.
        The first 7 rows of the CSV will be skipped during processing.
        """, unsafe_allow_html=True)

        with st.form("cost_report_form"):
            st.subheader("Input Parameters")
            col1, col2 = st.columns(2)
            
            with col1:
                uploaded_file = st.file_uploader(
                    "Upload CSV File",
                    type=["csv"],
                    help="Upload a CSV file containing service, monthly cost, and configuration summary columns. First 7 rows will be skipped."
                )
                usd_to_inr = st.number_input(
                    "USD to INR Exchange Rate",
                    min_value=0.0,
                    value=85.50,
                    step=0.01,
                    format="%.2f",
                    help="Enter the current USD to INR exchange rate."
                )
            
            with col2:
                customer_name = st.text_input(
                    "Customer Name",
                    value="",
                    help="Enter the customer name to be included in the report."
                )
                region = st.text_input(
                    "AWS Region",
                    value="US East (N. Virginia)",
                    help="Enter the AWS region for the workload."
                )
                pricing_link = st.text_input(
                    "Pricing Link",
                    value="",
                    help="Enter the pricing link to be included in the report."
                )
                output_filename = st.text_input(
                    "Output Excel File Name",
                    value=f"cost_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                    help="Specify the name for the output Excel file."
                )

            submit_button = st.form_submit_button("Generate Cost Report", type="primary")

        if submit_button:
            if not uploaded_file:
                st.error("Please upload a CSV file.")
            elif not output_filename.endswith('.xlsx'):
                st.error("Output file must have a .xlsx extension.")
            elif not customer_name.strip():
                st.error("Please enter a customer name.")
            else:
                with st.spinner("Generating cost report..."):
                    # Save uploaded file to temporary location
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".csv") as tmp_file:
                        tmp_file.write(uploaded_file.read())
                        tmp_file_path = tmp_file.name

                    # Generate output file path
                    output_file_path = os.path.join(tempfile.gettempdir(), output_filename)

                    # Initialize agent with user-provided usd_to_inr
                    agent = CostReportAgent(default_usd_to_inr=usd_to_inr)
                    result = agent.generate_cost_report(
                        input_file=tmp_file_path,
                        output_file=output_file_path,
                        customer_name=customer_name.strip(),
                        usd_to_inr=usd_to_inr,
                        region=region,
                        pricing_link=pricing_link.strip()
                    )

                    # Clean up temporary input file
                    try:
                        os.remove(tmp_file_path)
                    except Exception as e:
                        logger.warning("Failed to delete temporary file %s: %s", tmp_file_path, e)

                    if result["status"] == "success":
                        st.success("Cost report generated successfully!")
                        
                        # Display summary and AWS Services sheet
                        try:
                            df_summary = pd.read_excel(result["output_file"], sheet_name="Summary")
                            df_services = pd.read_excel(result["output_file"], sheet_name="AWS Services")
                            st.subheader("Cost Summary")
                            st.dataframe(df_summary, use_container_width=True)
                            st.subheader("AWS Services Details")
                            st.dataframe(df_services, use_container_width=True)
                        except Exception as e:
                            st.warning(f"Unable to display summary or services: {e}")

                        # Provide download link
                        with open(result["output_file"], "rb") as f:
                            st.download_button(
                                label="Download Cost Report",
                                data=f,
                                file_name=output_filename,
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                type="primary"
                            )
                    else:
                        st.error(f"Error generating cost report: {result['message']}")

                    # Clean up output file
                    try:
                        os.remove(output_file_path)
                    except Exception as e:
                        logger.warning("Failed to delete output file %s: %s", output_file_path, e)
        
        # Back button to return to landing page
        if st.button("Back to Home"):
            st.session_state.page = 'landing'

    main_pricing()

# Main function to handle page navigation
def main():
    if st.session_state.page == 'landing':
        landing_page()
    elif st.session_state.page == 'modernization':
        modernization_pathway()
    elif st.session_state.page == 'pricing':
        pricing_calculator()

if __name__ == "__main__":
    main()