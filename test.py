import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import re
import boto3
from botocore.exceptions import ClientError
import json
from datetime import datetime
import os
import logging
from typing import Dict, List, Optional

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

bedrock_client = boto3.client("bedrock-runtime", region_name="us-east-1")
LITE_MODEL_ID = "us.amazon.nova-lite-v1:0"


class CostReportAgent:
    def __init__(self, default_usd_to_inr: float, default_region: str = "US East (N. Virginia)"):
        self.usd_to_inr = default_usd_to_inr
        self.default_region = default_region
        logger.info("CostReportAgent initialized with USD to INR rate: %.2f", self.usd_to_inr)

    def extract_ec2_specs(self, instance_types: List[str]) -> Dict:
        if not instance_types:
            return {}

        system_prompt = [{"text": """
            You are a data extraction assistant specialized in AWS EC2 instances. 
            Given a list of EC2 instance types, provide the vCPUs and Memory (GiB) for each instance type.
            Return the result as a JSON object where keys are the instance types and values are dictionaries 
            with 'vCPUs' and 'MemoryGiB' keys. Use null if a value is not found.
            Ensure all specified instance types are included in the output, even if their values are not found.
        """}]

        user_message = [{"role": "user", "content": [{"text": f"Instance Types: {', '.join(instance_types)}"}]}]

        inference_params = {"maxTokens": 500, "topP": 0.9, "topK": 20, "temperature": 0.7}

        request_body = {
            "schemaVersion": "messages-v1",
            "messages": user_message,
            "system": system_prompt,
            "inferenceConfig": inference_params,
        }

        try:
            response = bedrock_client.invoke_model_with_response_stream(
                modelId=LITE_MODEL_ID, body=json.dumps(request_body)
            )
            full_response = ""
            for event in response.get("body", []):
                chunk = event.get("chunk")
                if chunk:
                    chunk_json = json.loads(chunk.get("bytes").decode())
                    if content_block_delta := chunk_json.get("contentBlockDelta"):
                        if text := content_block_delta.get("delta", {}).get("text"):
                            full_response += text

            json_start = full_response.find("```json") + 7
            json_end = full_response.rfind("```")
            if json_start > 6 and json_end > json_start:
                ec2_specs = json.loads(full_response[json_start:json_end].strip())
            else:
                ec2_specs = {}

            for it in instance_types:
                ec2_specs.setdefault(it, {"vCPUs": None, "MemoryGiB": None})

            return ec2_specs

        except Exception as e:
            logger.error(f"Error extracting EC2 specs: {e}")
            return {it: {"vCPUs": None, "MemoryGiB": None} for it in instance_types}

    def generate_service_description(self, service_name: str) -> str:
        system_prompt = [{"text": """
            You are an AI assistant specialized in AWS services. 
            Given the name of an AWS service, provide a highly precise, concise description (one sentence) 
            of its primary purpose in a cloud environment. Start with the service name, followed by a colon.
        """}]

        user_message = [{"role": "user", "content": [{"text": f"Service: {service_name}"}]}]

        inference_params = {"maxTokens": 80, "temperature": 0.4}

        request_body = {
            "schemaVersion": "messages-v1",
            "messages": user_message,
            "system": system_prompt,
            "inferenceConfig": inference_params,
        }

        try:
            response = bedrock_client.invoke_model_with_response_stream(
                modelId=LITE_MODEL_ID, body=json.dumps(request_body)
            )
            full_response = ""
            for event in response.get("body", []):
                chunk = event.get("chunk")
                if chunk:
                    chunk_json = json.loads(chunk.get("bytes").decode())
                    if content_block_delta := chunk_json.get("contentBlockDelta"):
                        if text := content_block_delta.get("delta", {}).get("text"):
                            full_response += text
            return full_response.strip()
        except Exception:
            return f"{service_name}: Provides core cloud functionality."

    def generate_best_practices(self, services: List[str]) -> List[str]:
        if not services:
            return ["No specific services detected. General AWS best practices apply."]

        services_list_str = ", ".join(services)

        system_prompt = [{"text": """
            You are an AWS cloud architect and cost/security optimization expert.
            Given a list of AWS services used in a customer's workload, generate 5 concise, actionable best practice recommendations.
            Focus on cost optimization, security, performance, and operational excellence.
            Number them 1 to 5.
            Make each recommendation specific to the services mentioned where possible.
            Keep each item short (1-2 sentences max).
            Start directly with the numbered list — no introduction text.
        """}]

        user_message = [{"role": "user", "content": [{"text": f"Services: {services_list_str}"}]}]

        inference_params = {"maxTokens": 400, "temperature": 0.6, "topP": 0.9}

        request_body = {
            "schemaVersion": "messages-v1",
            "messages": user_message,
            "system": system_prompt,
            "inferenceConfig": inference_params,
        }

        try:
            response = bedrock_client.invoke_model_with_response_stream(
                modelId=LITE_MODEL_ID, body=json.dumps(request_body)
            )
            full_response = ""
            for event in response.get("body", []):
                chunk = event.get("chunk")
                if chunk:
                    chunk_json = json.loads(chunk.get("bytes").decode())
                    if content_block_delta := chunk_json.get("contentBlockDelta"):
                        if text := content_block_delta.get("delta", {}).get("text"):
                            full_response += text

            lines = [line.strip() for line in full_response.split('\n') if line.strip() and re.match(r'^\d+\.', line.strip())]
            if lines:
                return lines[:5]
            else:
                return ["Model did not return numbered best practices."]

        except Exception as e:
            logger.error(f"Error generating best practices: {e}")
            return ["Unable to generate dynamic best practices at this time."]

    def extract_ec2_values(self, configuration_summary: str) -> tuple:
        try:
            os_match = re.search(r"operating\s*system\s*\((.*?)\)", configuration_summary, re.I)
            type_match = re.search(r"ec2\s*instance\s*\((.*?)\)", configuration_summary, re.I)
            price_match = re.search(r"pricing\s*strategy\s*\((.*?)\)", configuration_summary, re.I)

            return (
                os_match.group(1).strip() if os_match else None,
                type_match.group(1).strip() if type_match else None,
                price_match.group(1).strip() if price_match else None,
            )
        except:
            return None, None, None

    def generate_cost_report(
        self,
        input_file: str,
        output_file: str,
        customer_name: str,
        usd_to_inr: float,
        region: str,
        pricing_link: str = ""
    ):
        os.makedirs(os.path.dirname(output_file) or '.', exist_ok=True)

        header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        total_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
        pricing_link_fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        best_practice_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
        note_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        thin = Side(style='thin')
        full_border = Border(left=thin, right=thin, top=thin, bottom=thin)

        try:
            if not os.path.exists(input_file):
                raise FileNotFoundError(f"Input CSV not found: {input_file}")

            df = pd.read_csv(input_file, skiprows=7)
            df.columns = df.columns.str.lower().str.strip()

            service_col = next((c for c in df.columns if "service" in c), None)
            monthly_col = next((c for c in df.columns if "monthly" in c or "cost" in c), None)
            config_col = next((c for c in df.columns if "configuration" in c or "summary" in c or "config" in c), None)

            if not all([service_col, monthly_col, config_col]):
                raise ValueError("Required columns not found")

            data = df[[service_col, monthly_col, config_col]].fillna("")

            has_ec2 = any("EC2" in str(row[service_col]).upper() for _, row in data.iterrows())

            instance_types = set()
            ec2_specs = {}
            if has_ec2:
                for val in data[config_col]:
                    if "EC2" in str(val).upper():
                        m = re.search(r"ec2\s*instance\s*\((.*?)\)", str(val), re.I)
                        if m:
                            instance_types.add(m.group(1).strip())
                ec2_specs = self.extract_ec2_specs(list(instance_types))

            specs_failed = any(v["vCPUs"] is None and v["MemoryGiB"] is None for v in ec2_specs.values())

            # Collect unique cleaned service names for best practices & descriptions
            seen_services = set()
            cleaned_services = []
            for _, r in data.iterrows():
                svc = str(r[service_col]).strip()
                if svc and svc not in seen_services:
                    seen_services.add(svc)
                    clean_name = svc.split('(')[0].strip()
                    cleaned_services.append(clean_name)

            wb = openpyxl.Workbook()
            summary = wb.active
            summary.title = "Summary"

            summary.merge_cells("A1:B1")
            summary["A1"] = "Cost Estimation Summary"
            summary["A1"].font = Font(bold=True, color="000000")
            summary["A1"].alignment = Alignment(horizontal="center")
            summary["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            summary["A2"] = "Description"
            summary["B2"] = "Monthly Cost"
            for cell in [summary["A2"], summary["B2"]]:
                cell.fill = PatternFill(start_color="A7C5EB", end_color="A7C5EB", fill_type="solid")
                cell.border = full_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            total_usd = data[monthly_col].apply(pd.to_numeric, errors='coerce').sum()
            total_inr = total_usd * usd_to_inr

            summary["A3"] = "AWS Resource Cost (without TAX)"
            summary["B3"] = total_inr
            summary["B3"].number_format = "₹#,##0.00"
            summary["B3"].alignment = Alignment(horizontal="center")

            for r in [3]:
                for c in [1, 2]:
                    summary.cell(r, c).border = full_border

            summary.column_dimensions["A"].width = 30
            summary.column_dimensions["B"].width = 15

            # ────────────────────────────────
            #   AWS Services sheet
            # ────────────────────────────────

            sheet = wb.create_sheet("AWS Services")
            sheet.merge_cells("A1:K1" if has_ec2 else "A1:E1")
            sheet["A1"] = f"Cost Estimation Report For {customer_name}"
            sheet["A1"].font = Font(bold=True, color="000000")
            sheet["A1"].alignment = Alignment(horizontal="center")
            sheet["A1"].fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

            if has_ec2:
                headers = [
                    "S.NO", "EC2 Type (Advanced EC2 Instance)", "vCPU", "RAM", "Operating System",
                    "Running Hours", "Pricing Model", "Services", "Per Month USD", "Per Month INR", "Per Year INR"
                ]
                ec2_type_col = 2
                service_col_idx = 8
                usd_col = 9
                inr_col = 10
                yearly_col = 11
            else:
                headers = ["S.NO", "Services", "Per Month USD", "Per Month INR", "Per Year INR"]
                service_col_idx = 2
                usd_col = 3
                inr_col = 4
                yearly_col = 5

            for col, hdr in enumerate(headers, 1):
                cell = sheet.cell(2, col, hdr)
                cell.fill = header_fill
                cell.border = full_border
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            row = 3
            counter = 1

            rate_str = f"{usd_to_inr:.4f}"

            # ── Write data rows ──
            for _, r in data.iterrows():
                full_service = str(r[service_col]).strip()
                if not full_service:
                    continue
                try:
                    usd = float(r[monthly_col])
                except:
                    continue

                is_ec2 = "EC2" in full_service.upper()

                if has_ec2 and is_ec2:
                    os_val, ec2_type, price_model = self.extract_ec2_values(r[config_col])
                    spec = ec2_specs.get(ec2_type or "", {"vCPUs": None, "MemoryGiB": None})

                    # EC2 Type column gets only the instance type
                    values = [
                        (1, counter),
                        (2, ec2_type or ""),               # ← only type here
                        (3, spec.get('vCPUs')),
                        (4, spec.get('MemoryGiB')),
                        (5, os_val or ""),
                        (6, "730 hours"),
                        (7, price_model or ""),
                        (8, full_service),                 # ← full service name here
                        (9, usd)
                    ]

                    for col, val in values:
                        cell = sheet.cell(row, col, val)
                        cell.border = full_border
                        if col == 9:
                            cell.number_format = '$#,##0.00'
                        # Right align services & EC2 type columns
                        align_h = "right" if col in [2, 8] else ("left" if col < 9 else "right")
                        cell.alignment = Alignment(horizontal=align_h, vertical="center")

                    for col in [10, 11]:
                        formula = f"=I{row}*{rate_str}" if col == 10 else f"=J{row}*12"
                        cell = sheet.cell(row, col, value=formula)
                        cell.number_format = "₹#,##0.00"
                        cell.border = full_border
                        cell.alignment = Alignment(horizontal="right", vertical="center")

                else:
                    sheet.cell(row, 1, counter).border = full_border
                    sheet.cell(row, 1).alignment = Alignment(horizontal="right", vertical="center")

                    merge_end = usd_col - 1
                    sheet.merge_cells(f"B{row}:{openpyxl.utils.get_column_letter(merge_end)}{row}")
                    cell = sheet.cell(row, 2, full_service)
                    cell.alignment = Alignment(horizontal="right", vertical="center")  # right align
                    cell.border = full_border

                    for c in range(2, usd_col):
                        sheet.cell(row, c).border = full_border

                    sheet.cell(row, usd_col, usd).number_format = '$#,##0.00'
                    sheet.cell(row, usd_col).border = full_border
                    sheet.cell(row, usd_col).alignment = Alignment(horizontal="right", vertical="center")

                    sheet.cell(row, inr_col, value=f"={openpyxl.utils.get_column_letter(usd_col)}{row}*{rate_str}").number_format = "₹#,##0.00"
                    sheet.cell(row, inr_col).border = full_border
                    sheet.cell(row, inr_col).alignment = Alignment(horizontal="right", vertical="center")

                    sheet.cell(row, yearly_col, value=f"={openpyxl.utils.get_column_letter(inr_col)}{row}*12").number_format = "₹#,##0.00"
                    sheet.cell(row, yearly_col).border = full_border
                    sheet.cell(row, yearly_col).alignment = Alignment(horizontal="right", vertical="center")

                counter += 1
                row += 1

            last_data_row = row - 1

            # ── Total row ──
            total_r = row
            merge_end_total = "H" if has_ec2 else openpyxl.utils.get_column_letter(usd_col - 1)
            sheet.merge_cells(f"A{total_r}:{merge_end_total}{total_r}")
            sheet.cell(total_r, 1, "Total Cost").alignment = Alignment(horizontal="right")
            sheet.cell(total_r, 1).border = full_border

            for c in range(1, (12 if has_ec2 else usd_col + 3)):
                sheet.cell(total_r, c).border = full_border
                sheet.cell(total_r, c).fill = total_fill

            if last_data_row >= 3:
                sheet.cell(total_r, usd_col).value = f"=SUM({openpyxl.utils.get_column_letter(usd_col)}3:{openpyxl.utils.get_column_letter(usd_col)}{last_data_row})"
                sheet.cell(total_r, inr_col).value = f"=SUM({openpyxl.utils.get_column_letter(inr_col)}3:{openpyxl.utils.get_column_letter(inr_col)}{last_data_row})"
                sheet.cell(total_r, yearly_col).value = f"=SUM({openpyxl.utils.get_column_letter(yearly_col)}3:{openpyxl.utils.get_column_letter(yearly_col)}{last_data_row})"
            else:
                sheet.cell(total_r, usd_col).value = 0
                sheet.cell(total_r, inr_col).value = 0
                sheet.cell(total_r, yearly_col).value = 0

            sheet.cell(total_r, usd_col).number_format = '$#,##0.00'
            sheet.cell(total_r, inr_col).number_format = '₹#,##0.00'
            sheet.cell(total_r, yearly_col).number_format = '₹#,##0.00'

            # ── Pricing link ──
            pl_row = total_r + 1
            merge_end_pl = "H" if has_ec2 else openpyxl.utils.get_column_letter(usd_col - 1)
            sheet.merge_cells(f"A{pl_row}:{merge_end_pl}{pl_row}")
            sheet.cell(pl_row, 1, "Pricing Link").alignment = Alignment(horizontal="right")

            for c in range(1, (12 if has_ec2 else usd_col + 3)):
                sheet.cell(pl_row, c).border = full_border
                sheet.cell(pl_row, c).fill = pricing_link_fill

            sheet.cell(pl_row, usd_col if not has_ec2 else 9, pricing_link or "Not provided")

            # ── Notes section ── (NO borders)
            note_row = pl_row + 1

            note_title = sheet.cell(note_row, 1, "Note:")
            note_title.fill = note_fill

            notes = [
                "1. The given costs are considered as estimation based on the Monthly usage actual amount will get vary.",
                f"2. {region} region is considered for this workload.",
                f"3. The exchange rate is considered as ₹{usd_to_inr:.2f} as per {datetime.now().strftime('%d/%m/%y')} date."
            ]

            for i, note_text in enumerate(notes, 1):
                sheet.cell(note_row + i, 1, note_text)

            cur_note_row = note_row + len(notes) + 1
            note_sno = 4

            if specs_failed and has_ec2:
                sheet.cell(cur_note_row, 1, f"{note_sno}. Failed to extract EC2 specs (vCPU, RAM) from model.")
                cur_note_row += 1
                note_sno += 1

            # Service descriptions - no borders
            seen = set()
            for _, r in data.iterrows():
                svc = str(r[service_col]).strip()
                if svc and svc not in seen:
                    seen.add(svc)
                    clean = svc.split('(')[0].strip()
                    desc = self.generate_service_description(clean)
                    desc = desc.replace(f"{clean}:", "").strip()
                    sheet.cell(cur_note_row, 1, f"{note_sno}. {clean}: {desc}")
                    cur_note_row += 1
                    note_sno += 1

            # ── Best Practices - dynamically generated, no borders ──
            bp_row = cur_note_row
            bp_title = sheet.cell(bp_row, 1, "Best Practices:")
            bp_title.fill = best_practice_fill

            best_practices = self.generate_best_practices(cleaned_services)

            for i, line in enumerate(best_practices, 1):
                sheet.cell(bp_row + i, 1, line)

            # ── Column widths ──
            if has_ec2:
                widths = [6, 28, 10, 10, 18, 14, 16, 40, 14, 14, 14]
            else:
                widths = [6, 50, 16, 16, 16]

            for col, width in enumerate(widths, 1):
                sheet.column_dimensions[openpyxl.utils.get_column_letter(col)].width = width

            wb.save(output_file)
            logger.info(f"Report saved: {output_file}")
            return {"status": "success", "file": output_file}

        except Exception as e:
            logger.exception("Report generation failed")
            return {"status": "error", "message": str(e)}


def main():
    print("\n" + "="*70)
    print("        AWS Cost Report Generator (CSV → Excel)        ")
    print("="*70 + "\n")

    csv_path = input("1. Path to input CSV file : ").strip()
    if not os.path.exists(csv_path):
        print("Error: File not found!\n")
        return

    out_input = input("2. Output Excel file name (example: report) : ").strip()

    out_input = out_input.replace(".xlxs", "").replace(".xlsxx", "").replace(".xlsx", "").strip()

    if not out_input:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M")
        out_path = f"Cost_Report_{timestamp}.xlsx"
        print(f"→ No name entered → using: {out_path}")
    else:
        out_path = out_input

    if not out_path.lower().endswith('.xlsx'):
        out_path += ".xlsx"

    if not os.path.dirname(out_path):
        downloads = os.path.join(os.path.expanduser("~"), "Downloads")
        out_path = os.path.join(downloads, out_path)
        print(f"→ File will be saved to your Downloads folder: {out_path}")

    customer = input("3. Customer name : ").strip()
    if not customer:
        print("Error: Customer name is required!\n")
        return

    usd_str = input("4. USD to INR rate [default 85.50] : ").strip()
    usd_inr = 85.50 if not usd_str else float(usd_str)

    region = input("5. AWS Region [default: US East (N. Virginia)] : ").strip()
    region = region or "US East (N. Virginia)"

    link = input("6. Pricing link (optional) : ").strip()

    print("\n" + "-"*70)
    print("Generating report...")
    print(f"   Input   : {csv_path}")
    print(f"   Output  : {out_path}")
    print(f"   Customer: {customer}")
    print("-"*70 + "\n")

    agent = CostReportAgent(usd_inr, region)

    result = agent.generate_cost_report(
        input_file=csv_path,
        output_file=out_path,
        customer_name=customer,
        usd_to_inr=usd_inr,
        region=region,
        pricing_link=link
    )

    if result["status"] == "success":
        print("\n" + "═"*50)
        print("          REPORT GENERATED SUCCESSFULLY          ")
        print("═"*50)
        print(f"→ File saved to:\n  {result['file']}\n")
    else:
        print("\n" + "═"*50)
        print("               ERROR OCCURRED               ")
        print("═"*50)
        print(result["message"])
        print()


if __name__ == "__main__":
    main()